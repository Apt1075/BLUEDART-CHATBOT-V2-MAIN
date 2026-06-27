from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


EXPORT_DIR = Path("exports")
DOWNLOAD_PREFIX = "https://prompt.secutrak.in/api/v2/downloads"

TRIP_REPORT_HEADERS: List[str] = [
    "Route Category",
    "Route Type",
    "Region",
    "Origin",
    "Destination",
    "Route",
    "Route Sequence",
    "Fleet",
    "Trip ID",
    "Run Code",
    "Run Date",
    "Vehicle",
    "State",
    "Branch",
    "Area",
    "Driver Name",
    "Driver Number",
    "Priority",
    "Transporter",
    "STD",
    "ATD",
    "Delay Departure",
    "STA",
    "ATA",
    "TT-Mapped",
    "TT-Taken",
    "Delay Arrival",
    "Delay TT",
    "Schedule Halt",
    "Actual Halt",
    "ATT",
    "AHT",
    "GPS ATA",
    "Mobile ATA",
    "API ATA",
    "GPS ATD",
    "Mobile ATD",
    "API ATD",
    "Fixed GPS (Km)",
    "Fixed E-Lock (Km)",
    "Portable E-Lock (Km)",
    "Fixed GPS Exception",
    "Fixed ELock Exception",
    "Portable ELock Exception",
    "Supervisor Exception",
    "Status",
    "System Remarks",
    "Close By",
    "Close By Device",
    "Close Date",
    "Create By",
    "Total Bag",
    "Remark",
    "GPS Vendor",
    "Fixed E-Lock Vendor",
    "Portable E-Lock Vendor",
    "Primary Portable Device",
    "Secondary Portable Device",
    "Third Portable Device",
    "Fourth Portable Device",
]


def create_trip_report_excel(records: Iterable[Dict[str, Any]], stem: str = "trip_report") -> Dict[str, Any]:
    rows = [_to_report_row(record) for record in records]
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)

    filename = f"{_safe_filename(stem)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid4().hex[:8]}.xlsx"
    path = EXPORT_DIR / filename

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Trip Report"
    worksheet.append(TRIP_REPORT_HEADERS)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in rows:
        worksheet.append([row.get(header, "") for header in TRIP_REPORT_HEADERS])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    _autosize_columns(worksheet)
    workbook.save(path)

    return {
        "filename": filename,
        "path": str(path),
        "url": f"{DOWNLOAD_PREFIX}/{filename}",
        "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "row_count": len(rows),
    }


def _to_report_row(record: Dict[str, Any]) -> Dict[str, Any]:
    gps1 = _gps_exception(record.get("exception_common_backend"))
    gps2 = _gps_exception(record.get("exception_common_backend_2"))
    gps3 = _gps_exception(record.get("exception_common_backend_3"))
    trip_status = _trip_status(record.get("trip_status"))
    route_category = _route_category(record.get("trip_type"))
    atd = record.get("actual_source_departure_time") or ""
    ata = record.get("actual_destination_arrival_time") or ""
    close_remarks = record.get("close_remarks") or ""

    priority = "Standard"
    if record.get("high_shipment") == 1:
        priority = "High Shipment"

    return {
        "Route Category": route_category,
        "Route Type": record.get("shipment_method", ""),
        "Region": record.get("region_code", ""),
        "Origin": record.get("source_code", ""),
        "Destination": record.get("destination_code", ""),
        "Route": record.get("route_code", ""),
        "Route Sequence": record.get("route_name", ""),
        "Fleet": record.get("fleet_no", ""),
        "Trip ID": record.get("shipment_no", ""),
        "Run Code": record.get("run_code", ""),
        "Run Date": record.get("run_date", ""),
        "Vehicle": record.get("vehicle_no", ""),
        "State": record.get("state", ""),
        "Branch": record.get("branch_name", ""),
        "Area": record.get("area", ""),
        "Driver Name": record.get("driver_name", ""),
        "Driver Number": record.get("driver_mobile", ""),
        "Priority": priority,
        "Transporter": record.get("transporter_name", ""),
        "STD": record.get("schedule_departure", ""),
        "ATD": atd,
        "Delay Departure": record.get("delay_departure", ""),
        "STA": record.get("schedule_arrival", ""),
        "ATA": ata,
        "TT-Mapped": record.get("tt_mapped", ""),
        "TT-Taken": record.get("tt_taken", ""),
        "Delay Arrival": record.get("delay_arrival", ""),
        "Delay TT": record.get("delay_tt", "-"),
        "Schedule Halt": record.get("schedule_halt", ""),
        "Actual Halt": record.get("halt_duration", ""),
        "ATT": record.get("att", "-"),
        "PushTimeTrip": record.get("push_time_trip", ""),
        "PushStatus": record.get("push_status", "-"),
        "ServerGPSReceivedIn": record.get("server_gps_received_in", ""),
        "ServerGPSProcessedIn": record.get("server_gps_processed_in", ""),
        "ServerGPSReceivedOut": record.get("server_gps_received_out", ""),
        "ServerGPSProcessedOut": record.get("server_gps_processed_out", ""),
        "PushTimeIn": record.get("push_time_in", ""),
        "PushTimeOut": record.get("push_time_out", ""),
        "PushTimeInStatus": record.get("push_time_in_status", ""),
        "PushTimeOutStatus": record.get("push_time_out_status", ""),
        "Id": str(record.get("_id") or record.get("id") or ""),
        "RouteCategory": route_category,
        "HighShipment": record.get("high_shipment", 0),
        "Detail": _json_cell(record.get("detail", [])),
    }


def _gps_exception(value: Any) -> str:
    if value in (None, ""):
        return "GPS Active"
    return str(value)


def _trip_status(value: Any) -> str:
    try:
        return {0: "Completed", 1: "Schedule", 2: "Cancelled"}[int(value)]
    except (KeyError, TypeError, ValueError):
        return str(value or "")


def _route_category(value: Any) -> str:
    try:
        return {1: "Intercity", 2: "Intracity"}[int(value)]
    except (KeyError, TypeError, ValueError):
        return str(value or "")


def _json_cell(value: Any) -> str:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=True)
    return str(value or "")


def _safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_-]+", "_", value.strip().lower())
    return cleaned.strip("_") or "trip_report"


def _autosize_columns(worksheet: Any) -> None:
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column[:100]:
            max_length = max(max_length, len(str(cell.value or "")))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 36)
