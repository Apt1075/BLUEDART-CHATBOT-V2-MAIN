"""
generate_halt_report.py — Bluedart Halt Report Generator
=========================================================
Run: python generate_halt_report.py [hours]
Example: python generate_halt_report.py 3
"""

import httpx, asyncio, json, sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MONGO_API_URL = "http://3.110.228.31/secutrak_local_mongo/access/v0.1/selectQuery"
GROUP_ID      = "0041"

# KEY INSIGHT: trip_dashboard_live_status has pre-calculated fields:
#   stopped_gt_2h = 1  → vehicle stopped > 2 hours (use this to filter FIRST)
#   lastHaltTimeLR     → exact halt start time (use this to calculate exact duration)


def first(val):
    return val[0] if isinstance(val, list) and val else val


async def fetch_mongo(table, conditions, fields=None):
    fields  = fields or {}
    payload = {
        "conditions": json.dumps(conditions),
        "fields":     json.dumps(fields),
        "table":      table,
    }
    timeout = httpx.Timeout(connect=15.0, read=120.0, write=15.0, pool=15.0)
    for attempt in range(1, 4):
        try:
            print(f"   [{table}] attempt {attempt}...", end=" ", flush=True)
            async with httpx.AsyncClient(timeout=timeout) as client:
                resp = await client.post(
                    MONGO_API_URL, data=payload,
                    headers={"Content-Type": "application/x-www-form-urlencoded"},
                )
            if resp.status_code in (200, 201):
                data = resp.json()
                print(f"OK ({len(data) if isinstance(data, list) else 1} records)")
                return data
            print(f"HTTP {resp.status_code}")
            return []
        except httpx.ReadTimeout:
            print(f"timeout (attempt {attempt}/3)")
            if attempt == 3:
                return []
            await asyncio.sleep(3)
        except Exception as e:
            print(f"error: {e}")
            return []
    return []


def calc_halt(halt_time_str):
    """current_time - lastHaltTimeLR = exact halt duration"""
    if not halt_time_str:
        return 0, "N/A"
    try:
        halt_dt    = datetime.strptime(str(halt_time_str).strip(), "%Y-%m-%d %H:%M:%S")
        diff       = datetime.now() - halt_dt
        total_mins = max(0, int(diff.total_seconds() / 60))
        return total_mins, f"{total_mins//60}hrs {total_mins%60}min"
    except Exception:
        return 0, "N/A"


def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


async def generate_halt_report(threshold_hours=3):
    print(f"\n Generating {threshold_hours}+ Hours Halt Report...")
    print(f"   API: {MONGO_API_URL}\n")

    # ── STEP 1: Use stopped_gt_2h=1 to filter FIRST (much smaller result set)
    # This field is pre-calculated by Bluedart system — avoids full collection scan
    print("   Strategy: filter stopped_gt_2h=1 first, then calculate exact halt...")

    conditions = {
        "group_id":    GROUP_ID,
        "status":      1,
        "trip_status": 1,
        "stopped_gt_2h": 1,          # Pre-filtered by system — small result set
    }
    live_records = await fetch_mongo("trip_dashboard_live_status", conditions)

    # If stopped_gt_2h didn't work, try vehicle_status_current
    if not isinstance(live_records, list) or len(live_records) == 0:
        print("   Trying vehicle_status_current = Stopped fallback...")
        conditions2 = {
            "group_id":               GROUP_ID,
            "status":                 1,
            "trip_status":            1,
            "vehicle_status_current": "Stopped",
            "delay_hr":               {"$gte": 1},   # some delay — smaller set
        }
        live_records = await fetch_mongo("trip_dashboard_live_status", conditions2)

    if not isinstance(live_records, list) or len(live_records) == 0:
        print("\n   No data received.")
        print("   Check: API reachable but trip_dashboard_live_status may be slow.")
        print("   Try running during off-peak hours or contact DB team to add index.")
        return

    print(f"\n   Records with stopped_gt_2h=1: {len(live_records)}")

    # ── STEP 2: Calculate exact halt using lastHaltTimeLR ─────────────────────
    rows = []
    for rec in live_records:
        last_data = rec.get("last_data_current", {})

        # Get lastHaltTimeLR — most accurate halt start time
        halt_time = None
        if isinstance(last_data, dict):
            halt_time = first(last_data.get("lastHaltTimeLR"))
        if not halt_time:
            halt_time = rec.get("last_halt_time_current")
        if not halt_time:
            # Use stopped_duration field as fallback
            sd = str(rec.get("stopped_duration", ""))
            if ":" in sd:
                try:
                    parts = sd.split(":")
                    total_mins = int(parts[0])*60 + int(parts[1])
                    if total_mins >= threshold_hours * 60:
                        rows.append({
                            "shipment_no":     str(rec.get("shipment_no", "")),
                            "vehicle_no":      str(rec.get("vehicle_no", "")),
                            "run_date":        rec.get("create_date", ""),
                            "shipment_method": rec.get("shipment_method", ""),
                            "route_code":      "",
                            "hub_location":    "Yes" if rec.get("fixed_lock", 3) not in [3, None] else "No",
                            "from_date":       "calculated from stopped_duration",
                            "to_date":         datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "halt_duration":   f"{total_mins//60}hrs {total_mins%60}min",
                            "halt_mins":       total_mins,
                            "remarks":         f"Stopped {sd}",
                            "route_sequence":  "",
                            "location":        rec.get("last_address_current", ""),
                        })
                except Exception:
                    pass
            continue

        total_mins, halt_str = calc_halt(halt_time)

        # Apply user threshold
        if total_mins < threshold_hours * 60:
            continue

        rows.append({
            "shipment_no":     str(rec.get("shipment_no", "")),
            "vehicle_no":      str(rec.get("vehicle_no", "")),
            "run_date":        rec.get("create_date", ""),
            "shipment_method": rec.get("shipment_method", ""),
            "route_code":      "",
            "hub_location":    "Yes" if rec.get("fixed_lock", 3) not in [3, None] else "No",
            "from_date":       halt_time,
            "to_date":         datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "halt_duration":   halt_str,
            "halt_mins":       total_mins,
            "remarks":         f"Stopped Since {halt_time}",
            "route_sequence":  "",
            "location":        rec.get("last_address_current", ""),
        })

    rows.sort(key=lambda x: x["halt_mins"], reverse=True)
    print(f"   Vehicles > {threshold_hours}h halt: {len(rows)}")

    if not rows:
        print(f"\n   No vehicles halted more than {threshold_hours} hours found.")
        return

    # ── STEP 3: Enrich route info from courier_trip_detail ────────────────────
    shipment_nos = [r["shipment_no"] for r in rows if r["shipment_no"]]
    if shipment_nos:
        trip_data = await fetch_mongo(
            "courier_trip_detail",
            {"group_id": GROUP_ID, "status": 1, "shipment_no": {"$in": shipment_nos}},
        )
        if isinstance(trip_data, list):
            trip_map = {str(t.get("shipment_no")): t for t in trip_data}
            for row in rows:
                trip = trip_map.get(row["shipment_no"], {})
                if trip:
                    row["route_code"]      = trip.get("route_code", "")
                    row["route_sequence"]  = trip.get("route_name", "")
                    row["run_date"]        = trip.get("run_date", row["run_date"])
                    row["shipment_method"] = trip.get("shipment_method", row["shipment_method"])

    # ── STEP 4: Build Excel ───────────────────────────────────────────────────
    print(f"\n   Building Excel report...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    DARK_BLUE = "1F3864"
    MED_BLUE  = "2E75B6"
    LT_BLUE   = "D6E4F7"
    WHITE     = "FFFFFF"

    # Row 1: Title
    ws.merge_cells("A1:M1")
    ws["A1"] = f"VEHICLES HALTED FOR {int(threshold_hours)}+ HOURS"
    ws["A1"].font      = Font(bold=True, size=14, color=WHITE)
    ws["A1"].fill      = PatternFill("solid", fgColor=DARK_BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Row 2: Info
    ws.merge_cells("A2:M2")
    now_str = datetime.now().strftime("%m/%d/%Y, %I:%M:%S %p")
    ws["A2"] = (f"Generated on: {now_str}     |     "
                f"Total: {len(rows)} vehicles     |     "
                f"Threshold: {threshold_hours}+ hours     |     "
                f"Halt calculated from: lastHaltTimeLR")
    ws["A2"].font      = Font(italic=True, size=10, color="444444")
    ws["A2"].fill      = PatternFill("solid", fgColor=LT_BLUE)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    # Row 3: Headers
    headers    = ["S.No","Shipment No","Vehicle No","Run Date","Shipment Method",
                  "Route Code","Hub Location","From Date","To Date","Halt Duration",
                  "Remarks (Status)","Route Sequence","Location"]
    col_widths = [6, 14, 14, 22, 16, 14, 14, 22, 22, 14, 36, 30, 50]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font      = Font(bold=True, size=10, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor=MED_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 30

    # Data rows
    for ri, row in enumerate(rows, 1):
        r         = ri + 3
        halt_mins = row["halt_mins"]
        bg = ("FFD9D9" if halt_mins >= 1440 else
              "FFE5CC" if halt_mins >= 600  else
              "FFF2CC" if halt_mins >= 300  else
              "F2F2F2" if ri % 2 == 0 else "FFFFFF")

        fill = PatternFill("solid", fgColor=bg)
        vals = [ri, row["shipment_no"], row["vehicle_no"], row["run_date"],
                row["shipment_method"], row["route_code"], row["hub_location"],
                row["from_date"], row["to_date"], row["halt_duration"],
                row["remarks"], row["route_sequence"], row["location"]]

        for ci, val in enumerate(vals, 1):
            cell           = ws.cell(row=r, column=ci, value=val)
            cell.fill      = fill
            cell.border    = thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=(ci==13))
            cell.font      = Font(size=9)
            if ci == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if ci == 10:
                cell.font = Font(size=9, bold=True,
                    color="CC0000" if halt_mins>=600 else
                          "CC6600" if halt_mins>=300 else "000000")
        ws.row_dimensions[r].height = 18

    # Summary block
    critical = sum(1 for r in rows if r["halt_mins"] >= 1440)
    high     = sum(1 for r in rows if 600  <= r["halt_mins"] < 1440)
    medium   = sum(1 for r in rows if 300  <= r["halt_mins"] < 600)
    low      = sum(1 for r in rows if r["halt_mins"] < 300)
    sr       = len(rows) + 5

    ws.merge_cells(f"A{sr}:B{sr}")
    ws[f"A{sr}"] = "SUMMARY"
    ws[f"A{sr}"].font      = Font(bold=True, size=10, color=WHITE)
    ws[f"A{sr}"].fill      = PatternFill("solid", fgColor=DARK_BLUE)
    ws[f"A{sr}"].alignment = Alignment(horizontal="center")

    for i, (label, cnt, color) in enumerate([
        (f"Total > {int(threshold_hours)}h", len(rows), DARK_BLUE),
        ("Critical 24h+",  critical, "CC0000"),
        ("High 10-24h",    high,     "CC6600"),
        ("Medium 5-10h",   medium,   "996600"),
        (f"Low {int(threshold_hours)}-5h", low, "006600"),
    ]):
        lr = sr + 1 + i
        ws[f"A{lr}"] = label
        ws[f"B{lr}"] = cnt
        for col in "AB":
            ws[f"{col}{lr}"].font   = Font(size=9, bold=True, color=color)
            ws[f"{col}{lr}"].border = thin_border()
        ws[f"B{lr}"].alignment = Alignment(horizontal="center")

    # Legend
    lr0 = sr + 8
    ws[f"A{lr0}"] = "Color Legend:"
    ws[f"A{lr0}"].font = Font(bold=True, size=9)
    for i, (color, label) in enumerate([
        ("FFD9D9","24+ hrs — CRITICAL"), ("FFE5CC","10-24 hrs — HIGH"),
        ("FFF2CC","5-10 hrs — MEDIUM"),  ("F2F2F2","3-5 hrs — LOW"),
    ]):
        lr = lr0 + 1 + i
        ws[f"A{lr}"].fill   = PatternFill("solid", fgColor=color)
        ws[f"A{lr}"].border = thin_border()
        ws[f"B{lr}"]        = label
        ws[f"B{lr}"].font   = Font(size=9)

    ws.freeze_panes    = "A4"
    ws.auto_filter.ref = f"A3:M{len(rows)+3}"

    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"Halt_Report_{int(threshold_hours)}Plus_Hours_{timestamp}.xlsx"
    wb.save(output_file)

    print(f"\n Report: {output_file}")
    print(f"   Total={len(rows)} | Critical={critical} | High={high} | Medium={medium} | Low={low}")
    return output_file


if __name__ == "__main__":
    hours = float(sys.argv[1]) if len(sys.argv) > 1 else 3
    asyncio.run(generate_halt_report(hours))